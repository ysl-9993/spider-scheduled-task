# -*- coding: utf-8 -*-
"""
File Name: main.py
Author: yangshilin
Create Time: 2026-03-03 11:34
Description: 爬虫主程序
    调用api，需要预先登录公众号，登录地址：`API_BASE_URL`
    要抓取的公众号信息，写入target_accounts.json，fake_id从 `API_BASE_URL` 获取
"""

import requests
import json
import os
import datetime
from log_util import logger
from dotenv import load_dotenv
import time
import random
from openpyxl import Workbook, load_workbook
from lxml import etree
from datetime_util import timestamp_to_datetime_str

# 加载环境变量
load_dotenv()
# 输出根目录
output_base_dir = os.path.join(os.getenv("OUTPUT_BASE_PATH"))
# 列表json输出目录
output_list_dir = os.path.join(os.getenv("ARTICLE_LIST_PATH"))
# 文章html内容输出目录
output_content_dir = os.path.join(os.getenv("ARTICLE_CONTENT_PATH"))
# 最终xls输出目录
final_output_dir = os.path.join(os.getenv("FINAL_OUTPUT_PATH"))
# 定时任务间隔时间
gap_time = int(os.getenv("GAP_TIME"))
# 要抓取的公众号信息
target_accounts = json.load(open(os.getenv("TARGET_ACCOUNTS_PATH"), "r"))
# api基础url
api_base_url = os.getenv("API_BASE_URL")
# xls表头
xls_headers = json.load(open(os.getenv("XLS_HEADERS_PATH"), "r"))
# X-Auth-Key
key = os.getenv("X-Auth-Key")

# api接口
base_url = {
    # 获取文章列表api
    "get_article_list_url": api_base_url + "api/public/v1/article",
    # 下载文章内容api
    "download_url": api_base_url + "api/public/v1/download",
}

# 请求头 获取文章列表需要 若失效 从 `API_BASE_URL` 获取
headers = {
    "X-Auth-Key": key,
}

def init_path():
    """
    Description: 初始化输出目录结构
    Author: yangshilin
    Time: 2026-03-03 17:22
    """
    try:
        if not os.path.exists(output_base_dir):
            os.makedirs(output_base_dir, exist_ok=True)
        if not os.path.exists(output_list_dir):
            os.makedirs(output_list_dir, exist_ok=True)
        if not os.path.exists(output_content_dir):
            os.makedirs(output_content_dir, exist_ok=True)
        if not os.path.exists(final_output_dir):
            os.makedirs(final_output_dir, exist_ok=True)

        # 为每个公众号创建单独的目录    
        for account in target_accounts:
            # html文件输出目录
            if not os.path.exists(os.path.join(output_content_dir, account["name"])):
                os.makedirs(os.path.join(output_content_dir, account["name"]), exist_ok=True)
            # 文章列表json文件输出目录
            if not os.path.exists(os.path.join(output_list_dir, account["name"])):
                os.makedirs(os.path.join(output_list_dir, account["name"]), exist_ok=True)
            # xls文件输出目录
            if not os.path.exists(os.path.join(final_output_dir, account["name"])):
                os.makedirs(os.path.join(final_output_dir, account["name"]), exist_ok=True)
            if not os.path.exists(os.path.join(final_output_dir, account["name"], f"{account['name']}.xlsx")):
                # 创建xls文件并写入表头
                wb = Workbook()
                ws = wb.active
                ws.append(xls_headers)
                wb.save(os.path.join(final_output_dir, account["name"], f"{account['name']}.xlsx"))
                wb.close()
    except Exception as e:
        logger.error(f"错误信息：{e}")

def get_article_list():
    """
    Description: 获取文章列表
    Author: yangshilin
    Time: 2026-03-03 17:26
    Params: current_account: 当前公众号信息
    Returns: article_list: 文章列表
    """
    # 获取文章列表
    try:
        response = requests.get(base_url["get_article_list_url"] + "?fakeid=" + current_account["fake_id"], headers=headers)
        data = json.loads(response.text)
        # 提取文章列表
        article_list = data["articles"]
        # 当前时间字符串
        return article_list
    except Exception as e:
        logger.error(f"获取文章列表失败，错误信息：{e}，请检查X-Auth-Key是否过期")

def save_json(account_name, article_list):
    """
    Description: 保存json文件
    Author: yangshilin
    Time: 2026-03-03 17:28
    Params: account_name: 公众号名称
            article_list: 文章列表
    """
    try:
        current_time_str = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        # 保存文章列表json文件
        with open(os.path.join(output_list_dir, account_name, f"{current_time_str}.json"), "w") as f:
            json.dump(article_list, f, ensure_ascii=False, indent=4)
        logger.success(f"{account_name} 最新文章列表获取成功，已保存至 {os.path.join(output_list_dir, account_name, f'{current_time_str}.json')}")
    except Exception as e:
        logger.error(f"错误信息：{e}")

def filter_article(account_name, article_list):
    """
    Description: 过滤出新文章
    Author: yangshilin
    Time: 2026-03-03 17:45
    Params: account_name: 公众号名称
            article_list: 文章列表
    Returns: new_articles: 新文章列表
    """
    try:
        new_article_list = []
        xls_file_path = os.path.join(final_output_dir, account_name, f"{account_name}.xlsx")
        wb = load_workbook(xls_file_path, read_only=True)
        ws = wb.active
        # 获取表头
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        aid_col_index = None
        for idx, header in enumerate(header_row):
            if header == "文章id":
                aid_col_index = idx
                break
        # 提取已有的文章id
        xls_aids = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            aid = str(row[aid_col_index]).strip() if row[aid_col_index] else ""
            if aid and aid != "None":
                xls_aids.add(aid)
        # 过滤新文章
        for article in article_list:
            current_aid = article["aid"].strip()
            if current_aid not in xls_aids:
                new_article_list.append(article)
        wb.close()
        logger.info(f"{account_name} 发现 {len(new_article_list)} 篇新文章")
        return new_article_list
    except Exception as e:
        logger.error(f"错误信息：{e}")

def download_article(account_name, new_article_list):
    """
    Description: 下载新文章内容，更新xls文件
    Author: yangshilin
    Time: 2026-03-03 18:07
    Params: account_name: 公众号名称
            new_article_list: 新文章列表
    """
    for article in new_article_list:
        try:
            url = article["link"]
            # 下载文章不用携带请求头
            response = requests.get(base_url["download_url"] + "?url=" + url)
            html_data = response.text
            article["完整html内容"] = html_data
            # xpath解析纯文字内容
            tree = etree.HTML(html_data)
            word_data = tree.xpath(r'//*[@id="js_content"]/p/span/span/text()')
            # 合并所有段落文本
            article["纯文字内容"] = "\n".join(word_data)
            # 保存html文件
            with open(os.path.join(output_content_dir, account_name, article["title"] + ".html"), "w") as f:
                f.write(html_data)
            # 更新xls文件
            xls_file_path = os.path.join(final_output_dir, account_name, f"{account_name}.xlsx")
            wb = load_workbook(xls_file_path)
            ws = wb.active
            new_data_row = [
                article["aid"],
                article["title"],
                article["digest"],
                article["cover"],
                article["link"],
                timestamp_to_datetime_str(article["create_time"]),
                timestamp_to_datetime_str(article["update_time"]),
                article["author_name"],
                datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                article["完整html内容"],
                os.path.join(output_content_dir, account_name, article["title"] + ".html"),
                article["纯文字内容"],
            ]
            ws.append(new_data_row)
            wb.save(xls_file_path)
            wb.close()
            logger.success(f"{account_name} 新增文章 {article['title']} 成功")
            time.sleep(random.uniform(1, 3))
        except Exception as e:
            logger.error(f"错误信息：{e}")
    logger.success(f"{account_name}.xlsx 文件更新成功，新增 {len(new_article_list)} 条数据")

if __name__ == "__main__":
    # 初始化输出目录结构
    init_path()
    while True:
        for current_account in target_accounts:
            # 获取文章列表
            article_list = get_article_list()
            # 保存文章列表json文件
            save_json(current_account["name"], article_list)
            # 过滤出新文章
            new_article_list = filter_article(current_account["name"], article_list)
            # 下载新文章内容
            download_article(current_account["name"], new_article_list)
            # 冷却时间
            time.sleep(3)
        # 冷却时间
        current_time = datetime.datetime.now()
        next_time = current_time + datetime.timedelta(seconds=gap_time)
        next_time_str = next_time.strftime("%Y-%m-%d %H:%M:%S")
        logger.info(f"冷却中，{next_time_str} 进行下一次采集...")
        time.sleep(gap_time)
